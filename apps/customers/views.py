from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.db import transaction
from apps.accounts.models import User
from apps.tickets.models import Ticket
from .forms import CustomerCreationForm, CustomerEditForm, CustomerImportForm  # Added CustomerImportForm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import logging
import pandas as pd
import io
import random
import string
from datetime import datetime
from django.contrib.auth.hashers import make_password
import random
import string

logger = logging.getLogger(__name__)

@login_required
def customer_list(request):
    """List all customers with filtering and pagination."""
    if not request.user.is_admin():
        raise PermissionDenied("Only admins can view customer list.")
    
    # Base queryset
    customers = User.objects.filter(user_type='customer')
    
    # Get filter parameters
    search_query = request.GET.get('search', '')
    business_type = request.GET.get('business_type', '')
    status = request.GET.get('status', '')
    
    # Apply filters
    if search_query:
        customers = customers.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(business_name__icontains=search_query) |
            Q(phone_number__icontains=search_query)
        )
    
    if business_type:
        customers = customers.filter(business_type=business_type)
    
    if status == 'active':
        customers = customers.filter(is_active=True)
    elif status == 'inactive':
        customers = customers.filter(is_active=False)
    
    # Annotate with ticket statistics
    customers = customers.annotate(
        ticket_count=Count('created_tickets'),
        open_tickets=Count('created_tickets', 
            filter=Q(created_tickets__status__in=['new', 'open', 'in_progress', 'pending'])),
        resolved_tickets=Count('created_tickets',
            filter=Q(created_tickets__status='resolved'))
    ).order_by('-date_joined')
    
    # Pagination
    paginator = Paginator(customers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Business type choices for filter dropdown
    business_type_choices = User.BUSINESS_TYPE_CHOICES
    
    context = {
        'customers': page_obj,
        'page_obj': page_obj,
        'is_paginated': paginator.num_pages > 1,
        'search_query': search_query,
        'business_type': business_type,
        'status': status,
        'business_type_choices': business_type_choices,
    }
    
    return render(request, 'customers/customer_list.html', context)


@login_required
def customer_add(request):
    """Add a new customer."""
    if not request.user.is_admin():
        raise PermissionDenied("Only admins can add customers.")
    
    if request.method == 'POST':
        form = CustomerCreationForm(request.POST)
        if form.is_valid():
            customer = form.save()
            messages.success(request, f'Customer {customer.get_full_name()} created successfully.')
            return redirect('customers:customer_list')
    else:
        form = CustomerCreationForm()
    
    return render(request, 'customers/customer_form.html', {
        'form': form,
        'title': 'Add New Customer'
    })


@login_required
def customer_edit(request, pk):
    """Edit an existing customer."""
    if not request.user.is_admin():
        raise PermissionDenied("Only admins can edit customers.")
    
    customer = get_object_or_404(User, pk=pk, user_type='customer')
    
    if request.method == 'POST':
        form = CustomerEditForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, f'Customer {customer.get_full_name()} updated successfully.')
            return redirect('customers:customer_list')
    else:
        form = CustomerEditForm(instance=customer)
    
    return render(request, 'customers/customer_form.html', {
        'form': form,
        'customer': customer,
        'title': f'Edit Customer: {customer.get_full_name()}'
    })


@login_required
def customer_detail(request, pk):
    """View customer details and their tickets."""
    if not request.user.is_admin():
        raise PermissionDenied("Only admins can view customer details.")
    
    customer = get_object_or_404(User, pk=pk, user_type='customer')
    tickets = Ticket.objects.filter(created_by=customer).order_by('-created_at')
    
    # Ticket statistics
    total_tickets = tickets.count()
    open_tickets = tickets.filter(status__in=['new', 'open', 'in_progress', 'pending']).count()
    resolved_tickets = tickets.filter(status='resolved').count()
    closed_tickets = tickets.filter(status='closed').count()
    
    context = {
        'customer': customer,
        'tickets': tickets,
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
        'resolved_tickets': resolved_tickets,
        'closed_tickets': closed_tickets,
    }
    
    return render(request, 'customers/customer_detail.html', context)


@login_required
def customer_tickets(request, pk):
    """View all tickets for a specific customer."""
    if not request.user.is_admin():
        raise PermissionDenied("Only admins can view customer tickets.")
    
    customer = get_object_or_404(User, pk=pk, user_type='customer')
    tickets = Ticket.objects.filter(created_by=customer).order_by('-created_at')
    
    return render(request, 'customers/customer_tickets.html', {
        'customer': customer,
        'tickets': tickets
    })


@login_required
def customer_export(request):
    """Export customers to Excel."""
    if not request.user.is_admin():
        raise PermissionDenied("Only admins can export customer data.")
    
    # Get filter parameters (same as list view)
    search_query = request.GET.get('search', '')
    business_type = request.GET.get('business_type', '')
    status = request.GET.get('status', '')
    
    # Base queryset
    customers = User.objects.filter(user_type='customer')
    
    # Apply filters
    if search_query:
        customers = customers.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(business_name__icontains=search_query)
        )
    
    if business_type:
        customers = customers.filter(business_type=business_type)
    
    if status == 'active':
        customers = customers.filter(is_active=True)
    elif status == 'inactive':
        customers = customers.filter(is_active=False)
    
    # Annotate with ticket statistics
    customers = customers.annotate(
        ticket_count=Count('created_tickets'),
        open_tickets=Count('created_tickets', 
            filter=Q(created_tickets__status__in=['new', 'open', 'in_progress', 'pending'])),
        resolved_tickets=Count('created_tickets',
            filter=Q(created_tickets__status='resolved'))
    ).order_by('-date_joined')
    
    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Define headers
    headers = [
        'ID', 'Name', 'Email', 'Phone', 'Business Name', 'Business Type',
        'GST Number', 'Address', 'City', 'State', 'Country', 'Pincode',
        'Total Tickets', 'Open Tickets', 'Resolved Tickets', 'Status', 'Joined Date'
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data rows
    for row_num, customer in enumerate(customers, 2):
        ws.cell(row=row_num, column=1).value = customer.id
        ws.cell(row=row_num, column=2).value = customer.get_full_name()
        ws.cell(row=row_num, column=3).value = customer.email
        ws.cell(row=row_num, column=4).value = customer.phone_number or ''
        ws.cell(row=row_num, column=5).value = customer.business_name or ''
        ws.cell(row=row_num, column=6).value = customer.get_business_type_display() if customer.business_type else ''
        ws.cell(row=row_num, column=7).value = customer.gst_number or ''
        ws.cell(row=row_num, column=8).value = customer.address or ''
        ws.cell(row=row_num, column=9).value = customer.city or ''
        ws.cell(row=row_num, column=10).value = customer.state or ''
        ws.cell(row=row_num, column=11).value = customer.country or ''
        ws.cell(row=row_num, column=12).value = customer.pincode or ''
        ws.cell(row=row_num, column=13).value = customer.ticket_count or 0
        ws.cell(row=row_num, column=14).value = customer.open_tickets or 0
        ws.cell(row=row_num, column=15).value = customer.resolved_tickets or 0
        ws.cell(row=row_num, column=16).value = 'Active' if customer.is_active else 'Inactive'
        ws.cell(row=row_num, column=17).value = customer.date_joined.strftime('%Y-%m-%d')
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="customers_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    logger.info(f"Customers exported by {request.user.email}")
    return response


@login_required
@require_POST
def customer_toggle_status(request, pk):
    """Activate or deactivate a customer."""
    if not request.user.is_admin():
        raise PermissionDenied("Only admins can change customer status.")
    
    customer = get_object_or_404(User, pk=pk, user_type='customer')
    customer.is_active = not customer.is_active
    customer.save()
    
    status = "activated" if customer.is_active else "deactivated"
    messages.success(request, f'Customer {customer.get_full_name()} has been {status}.')
    
    return redirect('customers:customer_list')


@login_required
def customer_import(request):
    """Import customers from Excel file."""
    if not request.user.is_admin():
        raise PermissionDenied("Only admins can import customers.")
    
    if request.method == 'POST':
        form = CustomerImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            sheet_name = form.cleaned_data.get('sheet_name')
            has_header = form.cleaned_data.get('has_header', True)
            update_existing = form.cleaned_data.get('update_existing', False)
            
            try:
                # Read Excel file
                if sheet_name:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=0 if has_header else None)
                else:
                    df = pd.read_excel(excel_file, header=0 if has_header else None)
                
                # Process the data
                result = process_customer_import(df, update_existing, request.user)
                
                messages.success(
                    request, 
                    f"Import completed: {result['created']} created, {result['updated']} updated, "
                    f"{result['skipped']} skipped, {result['errors']} errors."
                )
                
                # Store detailed results in session for display
                request.session['import_results'] = result
                
                return redirect('customers:import_results')
                
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
                logger.error(f"Customer import error: {str(e)}")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = CustomerImportForm()
    
    # Get sample template for download
    context = {
        'form': form,
        'sample_fields': ['Name', 'Email', 'Phone', 'Business Name', 'Business Type', 
                          'GST Number', 'Address', 'City', 'State', 'Country', 'Pincode'],
    }
    return render(request, 'customers/customer_import.html', context)


def process_customer_import(df, update_existing=False, user=None):
    """
    Process the imported DataFrame and save to database.
    """
    result = {
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'errors': 0,
        'error_details': []
    }
    
    # Define field mappings (Excel column -> model field)
    field_mappings = {
        'Name': 'name',
        'Email': 'email',
        'Phone': 'phone_number',
        'Business Name': 'business_name',
        'Business Type': 'business_type',
        'GST Number': 'gst_number',
        'Address': 'address',
        'City': 'city',
        'State': 'state',
        'Country': 'country',
        'Pincode': 'pincode',
        'Status': 'is_active',
    }
    
    # Normalize column names (remove spaces, lowercase)
    df.columns = [str(col).strip() for col in df.columns]
    
    # Create a mapping dictionary for quick lookup
    column_map = {}
    for excel_col, model_field in field_mappings.items():
        # Find matching column (case-insensitive)
        for col in df.columns:
            if col.lower() == excel_col.lower():
                column_map[col] = model_field
                break
    
    if not column_map:
        raise ValueError("No matching columns found. Please ensure your Excel file has headers like: " + 
                        ", ".join(field_mappings.keys()))
    
    # Process each row
    for index, row in df.iterrows():
        try:
            with transaction.atomic():
                # Extract data from row
                data = {}
                for excel_col, model_field in column_map.items():
                    value = row[excel_col]
                    if pd.isna(value):
                        value = None
                    data[model_field] = value
                
                # Handle name splitting
                if 'name' in data and data['name']:
                    name_parts = str(data['name']).split(' ', 1)
                    first_name = name_parts[0]
                    last_name = name_parts[1] if len(name_parts) > 1 else ''
                else:
                    first_name = ''
                    last_name = ''
                
                # Handle business type (convert to choice value)
                business_type = data.get('business_type')
                if business_type:
                    # Convert to lowercase and match with choices
                    business_type = str(business_type).lower()
                    # Map common values to choice keys
                    business_type_map = {
                        'finance': 'finance',
                        'service': 'service',
                        'retail': 'retail',
                        'manufacturing': 'manufacturing',
                        'technology': 'technology',
                        'healthcare': 'healthcare',
                        'education': 'education',
                        'other': 'other',
                    }
                    business_type = business_type_map.get(business_type, 'other')
                
                # Handle status
                is_active = data.get('is_active', True)
                if isinstance(is_active, str):
                    is_active = is_active.lower() in ['active', 'yes', 'true', '1']
                
                # Check if customer exists
                email = data.get('email')
                if not email:
                    result['skipped'] += 1
                    result['error_details'].append(f"Row {index+2}: No email provided")
                    continue
                
                customer = User.objects.filter(email=email).first()
                
                if customer:
                    if update_existing:
                        # Update existing customer
                        customer.first_name = first_name
                        customer.last_name = last_name
                        customer.phone_number = data.get('phone_number')
                        customer.business_name = data.get('business_name')
                        customer.business_type = business_type
                        customer.gst_number = data.get('gst_number')
                        customer.address = data.get('address')
                        customer.city = data.get('city')
                        customer.state = data.get('state')
                        customer.country = data.get('country')
                        customer.pincode = data.get('pincode')
                        customer.is_active = is_active
                        customer.save()
                        result['updated'] += 1
                    else:
                        result['skipped'] += 1
                else:
                    # Create new customer
                    # Generate a random password for new customers
                    temp_password = ''.join(random.choices(string.ascii_letters + string.digits, k=12))
                    
                    customer = User.objects.create_user(
                        email=email,
                        password=temp_password,
                        first_name=first_name,
                        last_name=last_name,
                        user_type='customer',
                        phone_number=data.get('phone_number'),
                        business_name=data.get('business_name'),
                        business_type=business_type,
                        gst_number=data.get('gst_number'),
                        address=data.get('address'),
                        city=data.get('city'),
                        state=data.get('state'),
                        country=data.get('country'),
                        pincode=data.get('pincode'),
                        is_active=is_active
                    )
                    result['created'] += 1
                    
                    # Log the import
                    logger.info(f"Customer imported: {email} by {user.email if user else 'system'}")
                    
        except Exception as e:
            result['errors'] += 1
            result['error_details'].append(f"Row {index+2}: {str(e)}")
            logger.error(f"Error importing row {index+2}: {str(e)}")
    
    return result


@login_required
def import_results(request):
    """Display import results."""
    if not request.user.is_admin():
        raise PermissionDenied("Only admins can view import results.")
    
    results = request.session.get('import_results', {})
    return render(request, 'customers/import_results.html', {'results': results})


@login_required
def download_sample_template(request):
    """Download a sample Excel template for customer import."""
    if not request.user.is_admin():
        raise PermissionDenied("Only admins can download templates.")
    
    # Create a sample DataFrame
    data = {
        'Name': ['John Doe', 'Jane Smith'],
        'Email': ['john@example.com', 'jane@example.com'],
        'Phone': ['1234567890', '9876543210'],
        'Business Name': ['ABC Corp', 'XYZ Ltd'],
        'Business Type': ['Technology', 'Retail'],
        'GST Number': ['GST123456', 'GST789012'],
        'Address': ['123 Main St', '456 Oak Ave'],
        'City': ['New York', 'Los Angeles'],
        'State': ['NY', 'CA'],
        'Country': ['USA', 'USA'],
        'Pincode': ['10001', '90001'],
        'Status': ['Active', 'Active'],
    }
    
    df = pd.DataFrame(data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Customers', index=False)
    
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="customer_import_template.xlsx"'
    
    return response


@login_required
@require_POST
def bulk_customer_import(request):
    """
    Handle bulk import of customers from Excel file.
    This is an alternative endpoint for AJAX uploads.
    """
    if not request.user.is_admin():
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if 'file' not in request.FILES:
        return JsonResponse({'error': 'No file uploaded'}, status=400)
    
    excel_file = request.FILES['file']
    update_existing = request.POST.get('update_existing') == 'true'
    
    try:
        # Read Excel file
        df = pd.read_excel(excel_file)
        
        # Process the data
        result = process_customer_import(df, update_existing, request.user)
        
        return JsonResponse({
            'success': True,
            'created': result['created'],
            'updated': result['updated'],
            'skipped': result['skipped'],
            'errors': result['errors'],
            'error_details': result['error_details'][:10]  # Return first 10 errors
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
@login_required
def customer_reset_password(request, pk):
    """Reset customer password (admin only)."""
    if not request.user.is_admin():
        raise PermissionDenied("Only admins can reset customer passwords.")
    
    customer = get_object_or_404(User, pk=pk, user_type='customer')
    
    if request.method == 'POST':
        new_password = request.POST.get('new_password')
        if new_password:
            customer.set_password(new_password)
            customer.save()
            messages.success(request, f'Password for {customer.get_full_name()} has been reset successfully.')
            logger.info(f"Password reset for customer {customer.email} by admin {request.user.email}")
        else:
            messages.error(request, 'Please provide a new password.')
        return redirect('customers:customer_detail', pk=customer.pk)
    
    # Generate a random password suggestion
    suggested_password = ''.join(random.choices(string.ascii_letters + string.digits, k=12))
    
    return render(request, 'customers/customer_reset_password.html', {
        'customer': customer,
        'suggested_password': suggested_password
    })


@login_required
def customer_generate_random_password(request, pk):
    """Generate a random password for customer (AJAX)."""
    if not request.user.is_admin():
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    customer = get_object_or_404(User, pk=pk, user_type='customer')
    
    # Generate random password
    new_password = ''.join(random.choices(string.ascii_letters + string.digits, k=12))
    
    return JsonResponse({
        'success': True,
        'password': new_password
    })


@login_required
def customer_bulk_password_reset(request):
    """Reset passwords for multiple customers at once."""
    if not request.user.is_admin():
        raise PermissionDenied("Only admins can reset customer passwords.")
    
    if request.method == 'POST':
        customer_ids = request.POST.getlist('customer_ids')
        action = request.POST.get('action')
        
        if not customer_ids:
            messages.error(request, "No customers selected.")
            return redirect('customers:customer_list')
        
        customers = User.objects.filter(id__in=customer_ids, user_type='customer')
        
        if action == 'reset_to_default':
            # Reset to a default password
            default_password = 'Customer@123'
            for customer in customers:
                customer.set_password(default_password)
                customer.save()
            messages.success(request, f'Passwords reset to default for {customers.count()} customers.')
            
        elif action == 'generate_random':
            # Generate random passwords for each customer
            results = []
            for customer in customers:
                new_password = ''.join(random.choices(string.ascii_letters + string.digits, k=12))
                customer.set_password(new_password)
                customer.save()
                results.append({'email': customer.email, 'password': new_password})
            
            # Store results in session for display
            request.session['password_reset_results'] = results
            return redirect('customers:password_reset_results')
        
        return redirect('customers:customer_list')
    
    return redirect('customers:customer_list')


@login_required
def password_reset_results(request):
    """Display password reset results."""
    if not request.user.is_admin():
        raise PermissionDenied("Only admins can view password reset results.")
    
    results = request.session.get('password_reset_results', [])
    return render(request, 'customers/password_reset_results.html', {'results': results})