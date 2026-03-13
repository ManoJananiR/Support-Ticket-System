"""
Admin views for customer management.
"""

from django.shortcuts import render, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.http import HttpResponse
from .models import User
from .forms import AdminCustomerCreationForm, CustomerFilterForm
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


@staff_member_required
def admin_customer_list(request):
    """
    Admin view to list all customers with filters and export options.
    """
    # Base queryset - only customers
    customers = User.objects.filter(user_type='customer').order_by('-date_joined')
    
    # Apply filters
    form = CustomerFilterForm(request.GET or None)
    if form.is_valid():
        business_type = form.cleaned_data.get('business_type')
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        search = form.cleaned_data.get('search')
        
        if business_type:
            customers = customers.filter(business_type=business_type)
        
        if date_from:
            customers = customers.filter(date_joined__date__gte=date_from)
        
        if date_to:
            customers = customers.filter(date_joined__date__lte=date_to)
        
        if search:
            customers = customers.filter(
                models.Q(email__icontains=search) |
                models.Q(first_name__icontains=search) |
                models.Q(last_name__icontains=search) |
                models.Q(business_name__icontains=search) |
                models.Q(phone_number__icontains=search)
            )
    
    context = {
        'customers': customers,
        'form': form,
        'total_customers': customers.count(),
    }
    return render(request, 'admin/customer_list.html', context)


@staff_member_required
def admin_add_customer(request):
    """
    Admin view to add a new customer.
    """
    if request.method == 'POST':
        form = AdminCustomerCreationForm(request.POST)
        if form.is_valid():
            customer = form.save()
            messages.success(request, f'Customer {customer.get_full_name()} created successfully.')
            logger.info(f"Admin {request.user.email} created customer: {customer.email}")
            return redirect('admin_customer_list')
    else:
        form = AdminCustomerCreationForm()
    
    return render(request, 'admin/customer_form.html', {'form': form, 'is_create': True})


@staff_member_required
def admin_edit_customer(request, customer_id):
    """
    Admin view to edit an existing customer.
    """
    customer = User.objects.get(id=customer_id, user_type='customer')
    
    if request.method == 'POST':
        form = AdminCustomerEditForm(request.POST, instance=customer)
        if form.is_valid():
            customer = form.save()
            messages.success(request, f'Customer {customer.get_full_name()} updated successfully.')
            return redirect('admin_customer_list')
    else:
        form = AdminCustomerEditForm(instance=customer)
    
    return render(request, 'admin/customer_form.html', {'form': form, 'is_create': False})


@staff_member_required
def export_customers_excel(request):
    """
    Export customers to Excel file.
    """
    # Get filtered customers (same filters as list view)
    customers = User.objects.filter(user_type='customer').order_by('-date_joined')
    
    # Apply filters if present
    business_type = request.GET.get('business_type')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search')
    
    if business_type:
        customers = customers.filter(business_type=business_type)
    
    if date_from:
        customers = customers.filter(date_joined__date__gte=date_from)
    
    if date_to:
        customers = customers.filter(date_joined__date__lte=date_to)
    
    if search:
        customers = customers.filter(
            models.Q(email__icontains=search) |
            models.Q(first_name__icontains=search) |
            models.Q(last_name__icontains=search) |
            models.Q(business_name__icontains=search)
        )
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Define headers
    headers = [
        'ID', 'Email', 'First Name', 'Last Name', 'Phone Number',
        'Business Name', 'Business Type', 'GST Number',
        'Address', 'City', 'State', 'Country', 'Pincode',
        'Date Joined', 'Last Login', 'Active', 'Total Tickets'
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data
    for row_num, customer in enumerate(customers, 2):
        from apps.tickets.models import Ticket
        ticket_count = Ticket.objects.filter(created_by=customer).count()
        
        ws.cell(row=row_num, column=1).value = customer.id
        ws.cell(row=row_num, column=2).value = customer.email
        ws.cell(row=row_num, column=3).value = customer.first_name
        ws.cell(row=row_num, column=4).value = customer.last_name
        ws.cell(row=row_num, column=5).value = customer.phone_number
        ws.cell(row=row_num, column=6).value = customer.business_name
        ws.cell(row=row_num, column=7).value = customer.get_business_type_display() if customer.business_type else ''
        ws.cell(row=row_num, column=8).value = customer.gst_number
        ws.cell(row=row_num, column=9).value = customer.address
        ws.cell(row=row_num, column=10).value = customer.city
        ws.cell(row=row_num, column=11).value = customer.state
        ws.cell(row=row_num, column=12).value = customer.country
        ws.cell(row=row_num, column=13).value = customer.pincode
        ws.cell(row=row_num, column=14).value = customer.date_joined.strftime('%Y-%m-%d %H:%M')
        ws.cell(row=row_num, column=15).value = customer.last_login.strftime('%Y-%m-%d %H:%M') if customer.last_login else 'Never'
        ws.cell(row=row_num, column=16).value = 'Yes' if customer.is_active else 'No'
        ws.cell(row=row_num, column=17).value = ticket_count
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"customers_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    logger.info(f"Admin {request.user.email} exported customers to Excel")
    return response


@staff_member_required
def export_customers_csv(request):
    """
    Export customers to CSV file.
    """
    customers = User.objects.filter(user_type='customer').order_by('-date_joined')
    
    response = HttpResponse(content_type='text/csv')
    filename = f"customers_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Email', 'First Name', 'Last Name', 'Phone Number',
        'Business Name', 'Business Type', 'GST Number',
        'Address', 'City', 'State', 'Country', 'Pincode',
        'Date Joined', 'Last Login', 'Active'
    ])
    
    for customer in customers:
        from apps.tickets.models import Ticket
        writer.writerow([
            customer.id,
            customer.email,
            customer.first_name,
            customer.last_name,
            customer.phone_number or '',
            customer.business_name or '',
            customer.get_business_type_display() if customer.business_type else '',
            customer.gst_number or '',
            customer.address or '',
            customer.city or '',
            customer.state or '',
            customer.country or '',
            customer.pincode or '',
            customer.date_joined.strftime('%Y-%m-%d %H:%M'),
            customer.last_login.strftime('%Y-%m-%d %H:%M') if customer.last_login else 'Never',
            'Yes' if customer.is_active else 'No',
        ])
    
    logger.info(f"Admin {request.user.email} exported customers to CSV")
    return response